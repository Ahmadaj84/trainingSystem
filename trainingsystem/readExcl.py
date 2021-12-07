import openpyxl as oxl
from openpyxl import load_workbook
from flask import url_for
from trainingsystem import db, bcrypt
from trainingsystem.models import *


file_name = "C:\\Users\\ahmoo\\Documents\\GitHub\\trainingSystem\\trainingsystem\\test.xlsx"

#load the workbook

wb = load_workbook(file_name)
#get list of sheets in the workbook



sheet_list = wb.sheetnames
#loop over each sheet
db.create_all()

for sheet in sheet_list:
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2):
        user = []
        for cell in row:
            user.append(cell.value)
        hashed_password = bcrypt.generate_password_hash(user[3])
        db.session.add(User(username = user[0], uniID = user[1],email = user[2],password=hashed_password))
        db.session.commit()
